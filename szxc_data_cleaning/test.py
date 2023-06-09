from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from random import randint

# wb = load_workbook(filename='./tests/test_import.xlsx')
# ws = wb.worksheets
# for sheet in ws:
#     for col in sheet.iter_cols():
#         print(col)
#         for cell in col:
#             print(cell.value)

wb = Workbook()
# ws = wb.create_sheet('test', 0)
font = Font(name='宋体', color='CCCCCC', size=14, italic=True)
thin = Side(border_style='thin', color='000000')
border = Border(top=thin, left=thin, right=thin, bottom=thin)
fill = PatternFill('solid', fgColor='DDDDDD')
alignment = Alignment(horizontal='center', vertical='center')
ws = wb.active
ws.title = 'test'

for i in range(1, 100, 1):
    for j in range(1, i + 1):
        tCell = ws.cell(row=i, column=j)
        tCell.value = '{0}-{1}'.format(str(i), str(j))
        tCell.font = font
        tCell.border = border
        tCell.fill = PatternFill('solid', fgColor=hex(randint(0, int(0xFFFFFF) + 1))[2:].upper()[::-1].zfill(6)[::-1])
        tCell.alignment = alignment

wb.save('./tests/test_export.xlsx')
wb.close()

str(123)
print(1)
