from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment


class Styles:
    def __init__(self):
        thin = Side(border_style='thin', color='000000')
        self.font = Font(name='宋体', color='000000', size=14, italic=False)
        self.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        self.fill = PatternFill('solid', fgColor='FFFFFF')
        self.alignment = Alignment(horizontal='center', vertical='center')

    def set_font(self, font='宋体', color='000000', size=14, italic=False) -> Font:
        self.font = Font(name=font, color=color, size=size, italic=italic)
        return self.font

    def set_fill(self, color: str) -> PatternFill:
        self.fill = PatternFill('solid', fgColor=color)
        return self.fill


class Operator:
    def __init__(self):
        self.import_data = {}
        self.export_data = {}
        # self.__default_style = Styles()

    def import_excel(self, path: str) -> {'': [{}]}:
        wb = load_workbook(filename=path)
        for sheet in wb.worksheets:
            for col_index, col in enumerate(sheet.iter_cols()):
                li = []
                for row_index, cell in enumerate(col):
                    li.append({
                        'value': str(cell.value) if cell.value is not None else None,
                        'row': row_index + 1,
                        'col': col_index + 1,
                        'style': Styles()
                    })
                self.import_data[li[0]['value']] = li[:]
            break  # just read the first sheet
        return self.import_data

    def export_excel(self,
                     title='sheet 0',
                     filename='new excel.xlsx',
                     ):
        print('----------\t处理中\t----------')
        wb = Workbook()
        ws = wb.active
        ws.title = title
        for key in self.export_data.keys():
            for value in self.export_data[key]:
                cell = ws.cell(row=value['row'], column=value['col'])
                cell.value = value['value']
                cell.font = value['style'].font
                cell.border = value['style'].border
                cell.fill = value['style'].fill
                cell.alignment = value['style'].alignment

        if filename.find('./') == -1:
            filename = './' + filename
        if filename.rfind('.xlsx') == -1:
            filename = filename + '.xlsx'
        wb.save(filename)
        wb.close()
        print('输出为' + filename)
        print('----------\t输出完成\t----------')

# if __name__ == '__main__':
#     print(import_excel('./tests/test_import.xlsx'))
#     test = []
#     for i in range(1, 100, 1):
#         for j in range(1, i + 1):
#             test.append({
#                 'value': i+j,
#                 'row': i,
#                 'col': j
#             })
#     export_excel(test, filename='tests/test.xlsx')
